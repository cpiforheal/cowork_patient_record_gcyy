import sys, json
sys.stdout.reconfigure(encoding='utf-8')
from openpyxl import load_workbook

p = r'c:\Users\Administrator\.trae-cn\attachments\6a810cefc829f70b711f9b2e\5d8d184c-ee0d-494b-bfd0-f74e327729db_3ebf509d-8d24-4450-9f4b-ddb4d2201542_副本科室耗材使用明细 (1).xlsx'
wb = load_workbook(p, data_only=True)
for ws in wb.worksheets:
    print(f'=== SHEET: {ws.title} dims={ws.dimensions} ===')
    for row in ws.iter_rows(values_only=True):
        cells = ['' if c is None else str(c).strip() for c in row]
        if any(cells):
            print(' | '.join(cells))
