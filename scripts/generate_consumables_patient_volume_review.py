from __future__ import annotations

import re
from collections import Counter
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


SOURCE = Path(r"C:\Users\Administrator\Desktop\副本科室耗材使用明细 (1).xlsx")
OUTPUT = Path(r"C:\Users\Administrator\Desktop\耗材使用明细_患者量分类复核.xlsx")

ROLE_BY_DEPARTMENT = {
    "理疗室": ("治疗师", "治疗/处置完成量"),
    "检验科": ("检验技师", "检验项目完成量（待检验接口）"),
    "护理部": ("护士长", "住院患者日 / 护理执行量"),
    "中医科": ("中医诊室人员", "中医处置完成量"),
    "手术室": ("手术护士", "手术类型与完成量"),
    "麻醉室": ("麻醉护士", "麻醉/手术完成量"),
    "胃肠镜": ("内镜护士", "内镜检查完成量"),
    "检查室": ("检查技师", "检查完成量及初/复诊"),
    "后勤保洁": ("保洁班长", "固定运行基线；高峰增量待核定"),
    "西药房": ("药师", "发药处方/患者数（待发药接口）"),
    "收费室": ("收费员", "收费、医保结算笔数（待收费接口）"),
    "中药房": ("中药师", "中药处方数（待处方接口）"),
}

DIRECT_PATIENT = "患者量计量"
CONDITIONAL_PATIENT = "条件患者计量"
FIXED = "固定运行消耗"
PENDING = "待核定（非固定）"
ON_DEMAND = "按需申领"


def cell(value: object) -> str:
    return "" if value is None else str(value).strip()


def classify(department: str, scenario: str, item: str, usage: str, note: str) -> tuple[str, str, str]:
    text = f"{scenario} {item} {usage} {note}".replace(" ", "")
    if "按需" in text or "正常损耗" in text:
        return ON_DEMAND, "不纳入患者量公式；由岗位按需申领并记录用途。", "按需触发"

    # 区间、约数、伤口/手术差异等，不能在只有患者总量时可靠扣减。
    ambiguous = (
        re.search(r"\d+\s*[-~～至]\s*\d+", text)
        or any(token in text for token in ("大概", "约", "根据", "伤口", "筛查需", "手术总量", "患者多时", "至少"))
    )
    if ambiguous:
        return PENDING, "保留科室录入与实际领用；待补充单次定额、服务类型或分摊口径后再进入自动扣减。", "待科室确认"

    # 有明确周期、每日基线或更换周期的是固定运行，不和患者数直接相乘。
    if re.search(r"(?:每|/)?\d+(?:天|周|月)|每天|每日|一周|两周|三天|天一换|一换", text):
        return FIXED, "按运行周期计入固定消耗；不参与患者量自动扣减。", "周期/班次"

    # 明确“每人一份/每患者一次”且没有额外条件，允许作为患者量直接系数。
    if (
        re.search(r"(?:1|一)(?:个|支|张|片|包|套|副|份)?/(?:人|患者|病人)|(?:每人|每患者)(?:1|一)|单人单次", text)
        or (department not in ("手术室", "麻醉室", "胃肠镜") and re.search(r"(?:门诊|住院|新病号|复查).*?(?:患者|病患)", scenario))
    ):
        return DIRECT_PATIENT, "患者数 × 已确认单次用量，自动形成预估与扣减。", "来患者数量"

    return CONDITIONAL_PATIENT, "需要以检查、处置、处方、住院日或服务类型等业务完成量替代单纯来患者数后计算。", ROLE_BY_DEPARTMENT.get(department, ("科室人员", "业务完成量"))[1]


def source_rows() -> list[dict[str, str]]:
    workbook = load_workbook(SOURCE, data_only=True)
    records: list[dict[str, str]] = []
    for sheet in workbook.worksheets:
        headers = [cell(c.value) for c in sheet[1]]
        for row_index, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
            values = {headers[i]: cell(value) for i, value in enumerate(row) if i < len(headers)}
            item = values.get("耗材名称") or values.get("名称") or values.get("物品名称")
            if not item:
                continue
            scenario = values.get("适用场景/患者类型") or values.get("用途") or values.get("使用场景") or ""
            usage = values.get("使用数量") or values.get("用量") or values.get("使用量") or values.get("规格/用量") or ""
            specification = values.get("规格") or values.get("型号") or values.get("规格/用量") or ""
            purpose = scenario
            note = values.get("备注") or values.get("说明") or ""
            category, calculation, variable = classify(sheet.title, scenario, item, usage, f"{purpose} {note}")
            role, trigger = ROLE_BY_DEPARTMENT.get(sheet.title, ("科室人员", "待确认"))
            records.append({
                "科室": sheet.title,
                "来源行": str(row_index),
                "耗材名称": item,
                "规格": specification,
                "原始使用数量": usage,
                "用途/场景": purpose,
                "原始备注": note,
                "分类": category,
                "计算/处理方式": calculation,
                "数量变量": variable,
                "岗位录入人": role,
                "建议触发数据": trigger,
                "日预估": "由科室填写",
                "周预估": "日预估 × 排班天数（默认 7）",
                "确认状态": "待科室确认" if category in (PENDING, CONDITIONAL_PATIENT) else "待启用",
            })

    # 原表中的垃圾袋为“固定 46 个/天 + 高峰增量”，显式拆分高峰增量，避免被固定基线掩盖。
    records.append({
        "科室": "后勤保洁", "来源行": "2-增量", "耗材名称": "垃圾袋（患者高峰增量）", "规格": "同原表",
        "原始使用数量": "患者多时增加", "用途/场景": "高峰期垃圾收集", "原始备注": "固定基线 46 个/天仍按固定运行消耗维护",
        "分类": PENDING, "计算/处理方式": "先记录实际高峰增量；累计后确认每患者或每服务单的增量系数。",
        "数量变量": "高峰患者数 / 待确认", "岗位录入人": "保洁班长", "建议触发数据": "每日患者数与实际领用",
        "日预估": "待核定", "周预估": "待核定", "确认状态": "待科室确认",
    })
    return records


def write_sheet(workbook: Workbook, title: str, rows: list[dict[str, str]], columns: list[str]) -> None:
    sheet = workbook.create_sheet(title)
    sheet.append(columns)
    for record in rows:
        sheet.append([record.get(column, "") for column in columns])
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    for cell_ in sheet[1]:
        cell_.font = Font(bold=True, color="FFFFFF")
        cell_.fill = PatternFill("solid", fgColor="007F73")
        cell_.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for row in sheet.iter_rows(min_row=2):
        for cell_ in row:
            cell_.alignment = Alignment(vertical="top", wrap_text=True)
    for index, column in enumerate(columns, start=1):
        longest = max([len(column), *[len(str(row.get(column, ""))) for row in rows]], default=len(column))
        sheet.column_dimensions[get_column_letter(index)].width = min(max(longest + 2, 12), 42)


def main() -> None:
    records = source_rows()
    columns = ["科室", "来源行", "耗材名称", "规格", "原始使用数量", "用途/场景", "原始备注", "分类", "计算/处理方式", "数量变量", "岗位录入人", "建议触发数据", "日预估", "周预估", "确认状态"]
    workbook = Workbook()
    workbook.remove(workbook.active)
    counts = Counter(record["分类"] for record in records)
    overview = workbook.create_sheet("分类总览")
    overview.append(["患者量耗材重构口径", "说明"])
    overview.append(["总明细", f"原始 241 条 + 高峰增量拆分 1 条，共 {len(records)} 条。"])
    overview.append([DIRECT_PATIENT, "明确单次用量：患者数 × 单次定额，确认后可自动预估和扣减。"])
    overview.append([CONDITIONAL_PATIENT, "需检查、处置、处方、住院日等完成量；接口/口径确认后再扣减。"])
    overview.append([FIXED, "按天、周、月或班次运行；完全独立于患者数量。"])
    overview.append([PENDING, "定义模糊、区间/约数或缺少服务维度；已从自动公式剥离，保留科室实际录入。"])
    overview.append([ON_DEMAND, "按需领用或正常损耗；不纳入自动消耗。"])
    overview.append([])
    overview.append(["分类", "数量"])
    for category in (DIRECT_PATIENT, CONDITIONAL_PATIENT, FIXED, PENDING, ON_DEMAND):
        overview.append([category, counts[category]])
    write_sheet(workbook, "全部分类明细", records, columns)
    for category, title in ((DIRECT_PATIENT, "患者量计量"), (CONDITIONAL_PATIENT, "条件患者计量"), (FIXED, "固定运行消耗"), (PENDING, "待核定非固定"), (ON_DEMAND, "按需申领")):
        write_sheet(workbook, title, [row for row in records if row["分类"] == category], columns)

    department_rows = []
    for department in ROLE_BY_DEPARTMENT:
        department_records = [row for row in records if row["科室"] == department]
        role, trigger = ROLE_BY_DEPARTMENT[department]
        department_rows.append({
            "科室": department,
            "岗位入口": f"进销存系统 / 科室录入 / {department}",
            "责任岗位": role,
            "患者变量来源": trigger,
            "患者量计量": str(sum(row["分类"] == DIRECT_PATIENT for row in department_records)),
            "条件患者计量": str(sum(row["分类"] == CONDITIONAL_PATIENT for row in department_records)),
            "固定运行": str(sum(row["分类"] == FIXED for row in department_records)),
            "待核定非固定": str(sum(row["分类"] == PENDING for row in department_records)),
            "按需申领": str(sum(row["分类"] == ON_DEMAND for row in department_records)),
            "每日患者预估": "由科室填写",
            "每周患者预估": "由系统汇总/科室核对",
            "首要动作": "先确认待核定项，再确认患者量计量项的单次定额",
        })
    write_sheet(workbook, "科室入口与职责", department_rows, list(department_rows[0]))

    for sheet in workbook.worksheets:
        sheet.sheet_view.showGridLines = False
    overview.freeze_panes = "A2"
    overview.column_dimensions["A"].width = 24
    overview.column_dimensions["B"].width = 78
    for cell_ in overview[1]:
        cell_.font = Font(bold=True, color="FFFFFF")
        cell_.fill = PatternFill("solid", fgColor="007F73")
    workbook.save(OUTPUT)
    print(f"Wrote {OUTPUT} with {len(records)} rows.")


if __name__ == "__main__":
    main()
